import pandas as pd
import psycopg2
import os
from dotenv import load_dotenv
import matplotlib.pyplot as plt
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill

load_dotenv()

def get_db_connection():
    return psycopg2.connect(
        host=os.getenv('DB_HOST', 'localhost'),
        database=os.getenv('DB_NAME', 'postgres'),
        user=os.getenv('DB_USER', 'postgres'),
        password=os.getenv('DB_PASSWORD', ''),
        port=os.getenv('DB_PORT', '5432')
    )
def execute_query(query):
    with get_db_connection() as conn:
        return pd.read_sql_query(query, conn)
    
def create_plot(df, config):
    plt.figure(figsize=(10, 6))
    
    if config['type'] == 'pie':
        plt.pie(df[config['y']], labels=df[config['x']], autopct='%1.1f%%')
    elif config['type'] == 'bar':
        plt.bar(df[config['x']], df[config['y']])
        plt.xticks(rotation=45)
    elif config['type'] == 'barh':
        plt.barh(df[config['x']], df[config['y']])
    elif config['type'] == 'line':
        plt.plot(df[config['x']], df[config['y']], marker='o', linewidth=2)
    elif config['type'] == 'hist':
        plt.hist(df[config['x']], bins=30, alpha=0.7, edgecolor='black')
    elif config['type'] == 'scatter':
        plt.scatter(df[config['x']], df[config['y']], alpha=0.5)
    
    plt.title(config['title'], fontsize=14, fontweight='bold')
    plt.tight_layout()
    plt.savefig(config['filename'], dpi=100, bbox_inches='tight')
    plt.close()

def interactive_plot():
    query = """
        SELECT 
            DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
            c.customer_state,
            ROUND(SUM(oi.price), 2) AS total_sales
        FROM olist_orders o
        JOIN olist_order_items oi ON o.order_id = oi.order_id
        JOIN olist_customers c ON o.customer_id = c.customer_id
        WHERE o.order_purchase_timestamp IS NOT NULL
        GROUP BY month, c.customer_state
        ORDER BY month, total_sales DESC;
    """
    df = execute_query(query)
    df["month"] = df["month"].dt.strftime("%Y-%m")

    fig = px.bar(
        df,
        x="customer_state",
        y="total_sales",
        color="customer_state",
        animation_frame="month",
        animation_group="customer_state",
        title="Динамика продаж по штатам (с временным ползунком)",
        labels={"customer_state": "Штат", "total_sales": "Продажи"}
    )
    fig.update_layout(xaxis={'categoryorder':'total descending'})
    fig.show()

def export_to_excel(dataframes_dict, filename="exports/ecommerce_report.xlsx"):
    os.makedirs("exports", exist_ok=True)

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(filename)
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        total_rows += ws.max_row - 1
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        for col in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            if all(isinstance(ws.cell(row=row, column=col).value, (int, float, type(None)))
                   for row in range(2, ws.max_row + 1)):
                cell_range = f"{col_letter}2:{col_letter}{ws.max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="FFAA0000",
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                    end_type="max", end_color="FF00AA00"
                )
                ws.conditional_formatting.add(cell_range, rule)

    wb.save(filename)
    print(f"\nСоздан файл {os.path.basename(filename)}, {len(dataframes_dict)} листа, {total_rows} строк")

def main():
    queries_config = {
        'payment_analysis': {
            'query': """
                SELECT 
                    p.payment_type, 
                    COUNT(*) AS payment_count
                FROM olist_orders o
                JOIN olist_order_payments p ON o.order_id = p.order_id
                JOIN olist_customers c ON o.customer_id = c.customer_id
                GROUP BY p.payment_type
                ORDER BY payment_count DESC;
            """,
            'type': 'pie', 'x': 'payment_type', 'y': 'payment_count',
            'title': 'Распределение заказов по методам оплаты', 
            'filename': 'charts/payment_types_pie.png'
        },
        'top_categories': {
            'query': """
                SELECT 
                    pt.product_category_name_english AS category_name,
                    ROUND(SUM(oi.price), 2) AS total_revenue
                FROM olist_order_items oi
                JOIN olist_products p ON oi.product_id = p.product_id
                JOIN product_category_name_translation pt 
                    ON p.product_category_name = pt.product_category_name
                GROUP BY pt.product_category_name_english
                ORDER BY total_revenue DESC
                LIMIT 10;
            """,
            'type': 'bar', 'x': 'category_name', 'y': 'total_revenue',
            'title': 'Топ-10 категорий товаров по выручке', 
            'filename': 'charts/top_categories_bar.png'
        },
        'delivery_analysis': {
            'query': """
                SELECT 
                    c.customer_state,
                    ROUND(AVG(EXTRACT(EPOCH FROM (o.order_delivered_customer_date - o.order_purchase_timestamp)) / 86400), 2) AS avg_delivery_days
                FROM olist_orders o
                JOIN olist_customers c ON o.customer_id = c.customer_id
                WHERE o.order_status = 'delivered' 
                  AND o.order_delivered_customer_date IS NOT NULL
                GROUP BY c.customer_state
                HAVING COUNT(o.order_id) >= 50
                ORDER BY avg_delivery_days DESC;
            """,
            'type': 'barh', 'x': 'customer_state', 'y': 'avg_delivery_days',
            'title': 'Среднее время доставки по штатам', 
            'filename': 'charts/delivery_by_state_barh.png'
        },
        'monthly_sales': {
            'query': """
                SELECT 
                    DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
                    ROUND(SUM(oi.price), 2) AS total_sales
                FROM olist_orders o
                JOIN olist_order_items oi ON o.order_id = oi.order_id
                WHERE o.order_purchase_timestamp IS NOT NULL
                GROUP BY month
                ORDER BY month;
            """,
            'type': 'line', 'x': 'month', 'y': 'total_sales',
            'title': 'Динамика продаж по месяцам', 
            'filename': 'charts/monthly_sales_line.png'
        },
        'order_values': {
            'query': """
                SELECT 
                    p.payment_value
                FROM olist_order_payments p
                JOIN olist_orders o ON o.order_id = p.order_id
                WHERE p.payment_value IS NOT NULL;
            """,
            'type': 'hist', 'x': 'payment_value',
            'title': 'Распределение стоимости заказов', 
            'filename': 'charts/order_values_hist.png'
        },
        'freight_vs_payment': {
            'query': """
                SELECT 
                    oi.freight_value,
                    p.payment_value
                FROM olist_order_items oi
                JOIN olist_order_payments p ON oi.order_id = p.order_id
                JOIN olist_orders o ON o.order_id = oi.order_id
                WHERE oi.freight_value IS NOT NULL 
                  AND p.payment_value IS NOT NULL;
            """,
            'type': 'scatter', 'x': 'freight_value', 'y': 'payment_value',
            'title': 'Зависимость стоимости доставки от суммы заказа', 
            'filename': 'charts/freight_vs_payment_scatter.png'
        }
    }

    os.makedirs("charts", exist_ok=True)

    excel_data = {}

    for name, config in queries_config.items():
        df = execute_query(config['query'])
        if df is not None and not df.empty:
            create_plot(df, config)
            excel_data[name] = df
    orders_df = execute_query("""
        SELECT 
            o.order_id,
            o.order_status,
            o.order_purchase_timestamp,
            o.order_delivered_customer_date,
            c.customer_city,
            c.customer_state,
            p.payment_type,
            p.payment_value
        FROM olist_orders o
        JOIN olist_customers c ON o.customer_id = c.customer_id
        JOIN olist_order_payments p ON o.order_id = p.order_id
        LIMIT 1000;
    """)
    excel_data["orders_sample"] = orders_df

    export_to_excel(excel_data)

    print("\n Открываю интерактивный график...")
    interactive_plot()

if __name__ == "__main__":
    main()
